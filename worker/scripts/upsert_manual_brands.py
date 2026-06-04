"""Upsert past-rounds brand data from the manual monitoring spreadsheet.

Policy (confirmed with user):
  - Add-only: if a (round, kg) already has any round_brands rows, SKIP it
    entirely (do NOT overwrite scraped data).
  - '메리츠' → '메리츠화재' canonical mapping.
  - 'X' / 'x' / '-' / '' → treat as empty (no brand).

Usage:
  uv run --with openpyxl python scripts/upsert_manual_brands.py            # dry-run
  uv run --with openpyxl python scripts/upsert_manual_brands.py --apply    # commit
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from worker.db import connect
from worker.lib.canonical_brand import HOST_TO_BRAND

XLSX = Path(r"C:\Users\MADUP\Desktop\서칭뷰 신검 과거 운영 브랜드.xlsx")

# Spreadsheet raw name → canonical brand name.
BRAND_ALIAS = {"메리츠": "메리츠화재"}

# Lower-cased tokens that mean "no brand".
EMPTY_TOKENS = {"x", "-", "—", ""}

# Canonical brand → preferred business_name when inserting a NEW brand row.
# For brands whose host is already in HOST_TO_BRAND we use it; otherwise we
# park them under a `__manual__::<slug>` sentinel that the cleanup detector
# does not consider broken.
NEW_BRAND_HOST = {
    "교보생명": "www.kyobo.co.kr",          # already in HOST_TO_BRAND
    "교보라이프플래닛": "__manual__::lifeplanet",
    "신한EZ손해보험": "__manual__::shinhanezfire",
    "한화손해보험": "__manual__::hanwhasonhae",
}


def normalize_brand(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if s.lower() in EMPTY_TOKENS:
        return None
    return BRAND_ALIAS.get(s, s)


def read_sv(ws) -> dict[tuple[int, str], list[str]]:
    out: dict[tuple[int, str], list[str]] = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2 or not row or row[1] is None:
            continue
        _, rnd, kg, _vc, brand = row[:5]
        key = (int(rnd), str(kg).strip())
        b = normalize_brand(brand)
        if b and b not in out[key]:
            out[key].append(b)
    return out


def read_np(ws) -> dict[tuple[int, str], list[str]]:
    out: dict[tuple[int, str], list[str]] = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2 or not row or row[1] is None:
            continue
        _, rnd, _period, _l1, _l2, kg, _mb, b1, b2 = row[:9]
        key = (int(rnd), str(kg).strip())
        for b in (b1, b2):
            nb = normalize_brand(b)
            if nb and nb not in out[key]:
                out[key].append(nb)
    return out


def main(apply: bool) -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    sv = read_sv(wb["서칭뷰"])
    np = read_np(wb["신검"])

    canonical_hosts = set(HOST_TO_BRAND.keys())

    with connect() as conn:
        with conn.cursor() as cur:
            # Lookup tables
            cur.execute("SELECT id, code FROM products")
            prod = {code: pid for pid, code in cur.fetchall()}

            cur.execute(
                """
                SELECT rkg.id, r.product_id, r.round_no, kg.name
                FROM round_keyword_groups rkg
                JOIN rounds r ON r.id = rkg.round_id
                JOIN keyword_groups kg ON kg.id = rkg.keyword_group_id
                """
            )
            rkg_by_key = {(pid, rno, name): rkgid for rkgid, pid, rno, name in cur.fetchall()}

            cur.execute("SELECT DISTINCT round_keyword_group_id FROM round_brands")
            rkgs_with_data = {row[0] for row in cur.fetchall()}

            cur.execute(
                """
                SELECT b.id, b.business_name, b.display_name, COUNT(rb.id) AS uses
                FROM brands b
                LEFT JOIN round_brands rb ON rb.brand_id = b.id
                GROUP BY b.id, b.business_name, b.display_name
                """
            )
            brand_by_display: dict[str, list[tuple[int, str, int]]] = defaultdict(list)
            for bid, bn, dn, uses in cur.fetchall():
                brand_by_display[dn].append((bid, bn, uses))

            # Pick best brand_id per canonical display_name.
            brand_id_by_display: dict[str, int] = {}
            for dn, rows in brand_by_display.items():
                def rank(r: tuple[int, str, int]) -> tuple[int, int, int]:
                    bid, bn, uses = r
                    return (-int(bn in canonical_hosts), -uses, bid)
                best = sorted(rows, key=rank)[0]
                brand_id_by_display[dn] = best[0]

            # Insert any missing brands.
            for canonical_name, host in NEW_BRAND_HOST.items():
                if canonical_name in brand_id_by_display:
                    continue
                if not apply:
                    print(f"  [dry] would insert brand: {canonical_name!r} (host={host!r})")
                    brand_id_by_display[canonical_name] = -1  # placeholder
                    continue
                cur.execute(
                    """
                    INSERT INTO brands (business_name, display_name, aliases)
                    VALUES (%s, %s, '[]'::jsonb)
                    ON CONFLICT (business_name) DO UPDATE SET display_name = EXCLUDED.display_name
                    RETURNING id
                    """,
                    (host, canonical_name),
                )
                new_id = cur.fetchone()[0]
                brand_id_by_display[canonical_name] = new_id
                print(f"  [+] inserted brand: {canonical_name!r} id={new_id} host={host!r}")

            # Build the insert plan
            plan: list[tuple[int, int, int, str, str]] = []
            unresolved: dict[str, int] = defaultdict(int)
            skipped_existing = 0
            skipped_no_rkg = 0

            for label, prod_key, data in (
                ("SV", "SEARCHING_VIEW", sv),
                ("NP", "NEW_PRODUCT", np),
            ):
                pid = prod[prod_key]
                for (rnd, kg), brands in data.items():
                    rkg_id = rkg_by_key.get((pid, rnd, kg))
                    if rkg_id is None:
                        skipped_no_rkg += 1
                        continue
                    if rkg_id in rkgs_with_data:
                        skipped_existing += 1
                        continue
                    for slot, b in enumerate(brands, start=1):
                        bid = brand_id_by_display.get(b)
                        if bid is None or bid < 0:
                            unresolved[b] += 1
                            continue
                        plan.append((rkg_id, bid, slot, b, f"{label} r{rnd} {kg}"))

            print(f"\n=== Plan summary ===")
            print(f"  rows to insert         : {len(plan)}")
            print(f"  skipped (already data) : {skipped_existing}")
            print(f"  skipped (no rkg in DB) : {skipped_no_rkg}")
            if unresolved:
                print(f"  UNRESOLVED brands: {dict(unresolved)}")

            # Sample
            print(f"\n  first 10 planned rows:")
            for rkg_id, bid, slot, b, ctx in plan[:10]:
                print(f"    rkg={rkg_id} brand_id={bid} slot={slot} {b!r:<15} ({ctx})")

            if not apply:
                print("\n[dry-run] re-run with --apply to commit.")
                return

            # Apply
            inserted = 0
            skipped_conflict = 0
            for rkg_id, bid, slot, _b, _ctx in plan:
                cur.execute(
                    """
                    INSERT INTO round_brands (
                        round_keyword_group_id, brand_id, slot_no,
                        display_name, sub_title, description,
                        source, confidence
                    )
                    VALUES (%s, %s, %s, NULL, NULL, NULL, 'manual', 1.0)
                    ON CONFLICT (round_keyword_group_id, slot_no) DO NOTHING
                    """,
                    (rkg_id, bid, slot),
                )
                if cur.rowcount:
                    inserted += cur.rowcount
                else:
                    skipped_conflict += 1

        conn.commit()
        print(f"\n  Inserted {inserted} round_brands rows ({skipped_conflict} skipped via ON CONFLICT).")


if __name__ == "__main__":
    main(apply="--apply" in sys.argv)
