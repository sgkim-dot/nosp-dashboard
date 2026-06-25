"""Export 'brand cleanup needed' suspects to an Excel file for manual review.

Mirrors the dashboard's /brand-cleanup logic: surfaces brands whose host is
broken or whose display_name is "(미확인 브랜드)" — the 긴급정정 category.

Run:
    uv run python scripts/export_brand_cleanup.py
    uv run python scripts/export_brand_cleanup.py --all     # 검토 필요까지 포함
    uv run python scripts/export_brand_cleanup.py --out path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from worker.db import connect

_HOST_BROKEN_RE = re.compile(r"[\s()가-힣]|^__unverified__|주식회사|회사명")

_PRODUCT_LABEL = {
    "SEARCHING_VIEW": "서칭뷰",
    "NEW_PRODUCT": "신제품검색",
    "ANNIVERSARY": "기념일",
}


def is_host_broken(host: str | None) -> bool:
    if not host:
        return True
    return bool(_HOST_BROKEN_RE.search(host))


def fetch_suspects(conn) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
              b.id AS brand_id,
              b.display_name,
              b.business_name,
              COUNT(rb.id)::int AS uses,
              array_agg(DISTINCT rb.display_name)
                FILTER (WHERE rb.display_name IS NOT NULL) AS ad_copies,
              array_agg(DISTINCT rb.sub_title)
                FILTER (WHERE rb.sub_title IS NOT NULL) AS sub_titles,
              array_agg(DISTINCT p.code || '|' || kg.name)
                FILTER (WHERE kg.name IS NOT NULL) AS contexts
            FROM brands b
            LEFT JOIN round_brands rb ON rb.brand_id = b.id
            LEFT JOIN round_keyword_groups rkg ON rkg.id = rb.round_keyword_group_id
            LEFT JOIN keyword_groups kg ON kg.id = rkg.keyword_group_id
            LEFT JOIN products p ON p.id = kg.product_id
            GROUP BY b.id, b.display_name, b.business_name
            HAVING COUNT(rb.id) > 0
            ORDER BY COUNT(rb.id) DESC, b.id
            """
        )
        rows = cur.fetchall()

    suspects: list[dict] = []
    for brand_id, display_name, business_name, uses, ad_copies, sub_titles, contexts in rows:
        if not display_name or not business_name:
            continue
        display = display_name.strip()
        host = business_name.strip()
        reasons: list[str] = []
        if is_host_broken(host):
            reasons.append("호스트 깨짐")
        if display == "(미확인 브랜드)":
            reasons.append("미확인 브랜드")
        if not reasons:
            continue

        # de-dup + filter empties
        copies = [c for c in (ad_copies or []) if c]
        subs = [s for s in (sub_titles or []) if s]
        ctxs = []
        for c in contexts or []:
            try:
                product, kg = c.split("|", 1)
            except ValueError:
                continue
            ctxs.append((_PRODUCT_LABEL.get(product, product), kg))

        suspects.append(
            {
                "brand_id": brand_id,
                "display_name": display,
                "business_name": host,
                "uses": uses,
                "reasons": ", ".join(reasons),
                "ad_copies": copies,
                "sub_titles": subs,
                "contexts": ctxs,
                "host_broken": is_host_broken(host),
            }
        )

    return suspects


def url_for(host: str, broken: bool) -> str:
    if broken:
        return ""
    if host.startswith("http"):
        return host
    return f"https://{host}"


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "긴급정정"

    headers = [
        "brand_id",
        "현재 표시명",
        "URL / host",
        "URL 링크",
        "사용 횟수",
        "사유",
        "광고 카피 (최대 5개)",
        "서브타이틀 (최대 5개)",
        "키워드 컨텍스트 (최대 5개)",
        "정정할 브랜드명 (작성)",
        "메모",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="FFEAB308")  # amber-500
    header_font = Font(bold=True, color="FFFFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for r in rows:
        link = url_for(r["business_name"], r["host_broken"])
        ws.append(
            [
                r["brand_id"],
                r["display_name"],
                r["business_name"],
                link,
                r["uses"],
                r["reasons"],
                "\n".join(r["ad_copies"][:5]),
                "\n".join(r["sub_titles"][:5]),
                "\n".join(f"{p} · {kg}" for p, kg in r["contexts"][:5]),
                "",  # 정정할 브랜드명
                "",  # 메모
            ]
        )

    widths = [10, 28, 38, 38, 10, 18, 50, 50, 40, 22, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # wrap text on multi-line cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output xlsx path. Defaults to project root with timestamp.",
    )
    args = parser.parse_args()

    with connect() as conn:
        all_suspects = fetch_suspects(conn)

    critical = [
        s
        for s in all_suspects
        if "호스트 깨짐" in s["reasons"] or "미확인 브랜드" in s["reasons"]
    ]

    if not critical:
        print("긴급정정 케이스 없음.")
        return

    if args.out:
        out_path = args.out
    else:
        kst = timezone(timedelta(hours=9))
        ts = datetime.now(kst).strftime("%Y%m%d_%H%M")
        out_path = (
            Path(__file__).resolve().parents[2]
            / f"브랜드정리_긴급정정_{ts}.xlsx"
        )

    write_xlsx(critical, out_path)
    print(f"긴급정정 {len(critical)}건 → {out_path}")


if __name__ == "__main__":
    main()
