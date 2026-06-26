"""Export ALL active-round brands for full audit.

Differs from export_brand_cleanup.py:
- That one exports only the 긴급정정 cases (sentinel host / 미확인 브랜드).
- This one exports EVERY brand that's actually used this round so the operator
  can scan for cases where a raw English domain ended up as the brand name.

Column layout intentionally matches export_brand_cleanup.py so the existing
apply_brand_cleanup_xlsx.py works without modification — fill column J
(정정할 브랜드명), then run:
    uv run python scripts/apply_brand_cleanup_xlsx.py --file <this.xlsx> --apply

Run:
    uv run python scripts/export_brand_audit.py
    uv run python scripts/export_brand_audit.py --out path/to/file.xlsx
    uv run python scripts/export_brand_audit.py --all-rounds   # 활성 round 외에 전체
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from worker.db import connect


_PRODUCT_LABEL = {
    "SEARCHING_VIEW": "서칭뷰",
    "NEW_PRODUCT": "신제품검색",
    "ANNIVERSARY": "기념일",
}

# Heuristic: host that looks like a raw English domain (no Korean, looks like
# foo.com / foo.co.kr / smartstore.naver.com/X) AND display_name that's also
# raw-domain-ish (no Korean, contains '.com' or similar). These are the prime
# "raw domain became brand name" suspects.
_HAS_KOREAN = re.compile(r"[가-힣]")
_LOOKS_LIKE_DOMAIN = re.compile(r"\.[a-z]{2,6}(/|$)|^[a-z0-9.-]+\.[a-z]{2,6}", re.IGNORECASE)


def is_suspicious_raw_domain(display_name: str, host: str) -> bool:
    """display_name 자체가 도메인 모양이면 의심."""
    if _HAS_KOREAN.search(display_name):
        return False
    if _LOOKS_LIKE_DOMAIN.search(display_name):
        return True
    return False


def fetch_all_brands(conn, active_only: bool) -> list[dict]:
    where_clause = ""
    if active_only:
        where_clause = """
            AND EXISTS (
                SELECT 1 FROM rounds r
                WHERE r.id = rkg.round_id
                  AND r.period_start <= (NOW() AT TIME ZONE 'Asia/Seoul')::date
                  AND r.period_end   >= (NOW() AT TIME ZONE 'Asia/Seoul')::date
            )
        """
    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT
              b.id AS brand_id,
              b.display_name,
              b.business_name,
              COUNT(rb.id)::int AS uses,
              array_agg(DISTINCT rb.display_name)
                FILTER (WHERE rb.display_name IS NOT NULL) AS ad_copies,
              array_agg(DISTINCT rb.sub_title)
                FILTER (WHERE rb.sub_title IS NOT NULL) AS sub_titles,
              array_agg(DISTINCT p.code || '|' || kg.name)
                FILTER (WHERE kg.name IS NOT NULL) AS contexts
            FROM brands b
            JOIN round_brands rb ON rb.brand_id = b.id
            JOIN round_keyword_groups rkg ON rkg.id = rb.round_keyword_group_id
            LEFT JOIN keyword_groups kg ON kg.id = rkg.keyword_group_id
            LEFT JOIN products p ON p.id = kg.product_id
            WHERE 1=1
              {where_clause}
            GROUP BY b.id, b.display_name, b.business_name
            HAVING COUNT(rb.id) > 0
            ORDER BY COUNT(rb.id) DESC, b.id
            """
        )
        rows = cur.fetchall()

    out: list[dict] = []
    for brand_id, display_name, business_name, uses, ad_copies, sub_titles, contexts in rows:
        if not display_name or not business_name:
            continue
        display = display_name.strip()
        host = business_name.strip()

        # de-dup + filter empties
        copies = [c for c in (ad_copies or []) if c]
        subs = [s for s in (sub_titles or []) if s]
        ctxs = []
        for c in contexts or []:
            try:
                product, kg = c.split("|", 1)
            except ValueError:
                continue
            ctxs.append((_PRODUCT_LABEL.get(product, product), kg))

        flags = []
        if is_suspicious_raw_domain(display, host):
            flags.append("⚠ 영문도메인")
        if host.startswith("__unverified__"):
            flags.append("sentinel")
        if display == "(미확인 브랜드)":
            flags.append("미확인")

        out.append(
            {
                "brand_id": brand_id,
                "display_name": display,
                "business_name": host,
                "uses": uses,
                "flags": ", ".join(flags),
                "ad_copies": copies,
                "sub_titles": subs,
                "contexts": ctxs,
                "host_broken": host.startswith("__unverified__"),
            }
        )

    return out


def url_for(host: str, broken: bool) -> str:
    if broken:
        return ""
    if host.startswith("http"):
        return host
    return f"https://{host}"


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "브랜드_전수조사"

    headers = [
        "brand_id",
        "현재 브랜드명",
        "도메인 호스트",
        "URL 링크",
        "사용 횟수",
        "이슈 플래그",
        "광고 카피 (최대 5개)",
        "서브타이틀 (최대 5개)",
        "키워드 컨텍스트 (최대 5개)",
        "변경 브랜드명 (작성)",
        "메모",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="FF10B981")  # emerald-500
    header_font = Font(bold=True, color="FFFFFFFF")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    warn_fill = PatternFill("solid", fgColor="FFFEF3C7")  # amber-100 (suspicious row)

    for r in rows:
        link = url_for(r["business_name"], r["host_broken"])
        ws.append(
            [
                r["brand_id"],
                r["display_name"],
                r["business_name"],
                link,
                r["uses"],
                r["flags"],
                "\n".join(r["ad_copies"][:5]),
                "\n".join(r["sub_titles"][:5]),
                "\n".join(f"{p} · {kg}" for p, kg in r["contexts"][:5]),
                "",
                "",
            ]
        )
        # Highlight rows where display_name looks like a raw English domain
        if "영문도메인" in r["flags"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = warn_fill

    widths = [10, 32, 38, 38, 10, 20, 50, 50, 40, 24, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=None)
    parser.add_argument(
        "--all-rounds",
        action="store_true",
        help="기본은 활성 round만. 이 플래그를 주면 모든 round의 브랜드 포함.",
    )
    args = parser.parse_args()

    with connect() as conn:
        rows = fetch_all_brands(conn, active_only=not args.all_rounds)

    if not rows:
        print("브랜드 없음.")
        return

    if args.out:
        out_path = args.out
    else:
        kst = timezone(timedelta(hours=9))
        ts = datetime.now(kst).strftime("%Y%m%d_%H%M")
        scope = "전체" if args.all_rounds else "활성"
        out_path = (
            Path(__file__).resolve().parents[2]
            / f"브랜드_전수조사_{scope}_{ts}.xlsx"
        )

    write_xlsx(rows, out_path)
    n_warn = sum(1 for r in rows if "영문도메인" in r["flags"])
    n_sentinel = sum(1 for r in rows if "sentinel" in r["flags"])
    print(f"브랜드 {len(rows)}건 -> {out_path}")
    print(f"  [WARN] 영문도메인 의심: {n_warn}건 (노란색 강조)")
    print(f"  sentinel: {n_sentinel}건")
    print(f"\n작성 후 적용:")
    print(f'  uv run python scripts/apply_brand_cleanup_xlsx.py --file "{out_path}" --apply')


if __name__ == "__main__":
    main()
