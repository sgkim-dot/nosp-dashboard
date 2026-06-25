"""For each remaining brand-cleanup row, check whether the user's xlsx
mapped this exact host. If yes — it's a NEW row created after we applied
mappings (cycle 3 BAT still running). If no — the brand belongs to a
user-mapped advertiser but Naver hydrated a NEW ad-copy variant whose
sentinel host is different from any we registered.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from worker.db import connect
from worker.lib.canonical_brand import HOST_TO_BRAND

_HOST_BROKEN_RE = re.compile(r"[\s()가-힣]|^__unverified__|주식회사|회사명")
ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "브랜드정리_긴급정정_20260625_1039.xlsx"


def load_user_mappings() -> dict[str, str]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        host = row[2]
        fix = row[9] if len(row) > 9 else None
        if not (host and fix):
            continue
        out[str(host).strip()] = str(fix).strip()
    return out


def main() -> None:
    user_map = load_user_mappings()
    print(f"엑셀 매핑 {len(user_map)}건 로드")

    with connect() as c, c.cursor() as cur:
        cur.execute(
            """
            SELECT b.id, b.display_name, b.business_name,
                   COUNT(rb.id),
                   array_agg(DISTINCT rb.display_name) FILTER (WHERE rb.display_name IS NOT NULL)
            FROM brands b JOIN round_brands rb ON rb.brand_id=b.id
            GROUP BY b.id HAVING COUNT(rb.id) > 0
            """
        )
        rows = cur.fetchall()

    in_xlsx_exact: list = []
    in_canon_exact: list = []
    new_variants_of_mapped: list = []
    truly_new: list = []

    # Index xlsx by stripped sentinel content (the part after __unverified__::)
    def _strip(h: str) -> str:
        return h[len("__unverified__::"):] if h.startswith("__unverified__::") else h

    xlsx_keys_normalized = {_strip(k): v for k, v in user_map.items()}

    for bid, dn, bn, cnt, copies in rows:
        if not bn or not dn:
            continue
        if not _HOST_BROKEN_RE.search(bn) and dn != "(미확인 브랜드)":
            continue

        # 1. exact key in canonical_brand.HOST_TO_BRAND?
        if bn in HOST_TO_BRAND:
            in_canon_exact.append((bid, dn, bn, HOST_TO_BRAND[bn], copies))
            continue
        # 2. exact key in user xlsx (but not in canonical yet)?
        if bn in user_map:
            in_xlsx_exact.append((bid, dn, bn, user_map[bn], copies))
            continue
        # 3. NEW ad-copy variant of a brand the user did map?
        #    look for any user key whose stripped sentinel string is a prefix
        #    or shares the first 2-3 words with this brand's sentinel
        body = _strip(bn)
        match = None
        for k_body, brand in xlsx_keys_normalized.items():
            # share first 5 chars or longest-common-prefix > 6 → same family
            if body[:8] == k_body[:8] and len(body) >= 8:
                match = (k_body, brand)
                break
            # or share first word
            w1 = body.split(" ", 1)[0]
            kw1 = k_body.split(" ", 1)[0]
            if w1 and w1 == kw1:
                match = (k_body, brand)
                break
        if match:
            new_variants_of_mapped.append((bid, dn, bn, match[1], match[0], copies))
            continue
        truly_new.append((bid, dn, bn, copies))

    print(f"\n=== 분류 ===")
    print(f"  A. canonical에 이미 등록된 host인데 reconcile 안 된 것: {len(in_canon_exact)}")
    print(f"  B. xlsx 매핑은 있는데 canonical에는 없음 (적용 실패?): {len(in_xlsx_exact)}")
    print(f"  C. 사용자 매핑된 광고주의 새 카피 변형 (cycle 3가 새로 잡음): {len(new_variants_of_mapped)}")
    print(f"  D. 완전히 새로운 광고주 (매핑 없음): {len(truly_new)}")

    if in_canon_exact:
        print("\n[A] canonical 등록됐는데 brand 행이 살아있음 (reconcile 누락):")
        for bid, dn, bn, target, copies in in_canon_exact:
            print(f"  id={bid:5d}  dn={dn[:20]!r:22s}  → 정정: {target}")
            print(f"    host={bn[:90]}")
    if in_xlsx_exact:
        print("\n[B] xlsx 키와 정확히 일치하는데 canonical에 안 들어감 (버그):")
        for bid, dn, bn, target, copies in in_xlsx_exact:
            print(f"  id={bid:5d}  dn={dn[:20]!r:22s}  → 정정: {target}")
            print(f"    host={bn[:90]}")
    if new_variants_of_mapped:
        print("\n[C] 같은 광고주의 새 카피 변형 (cycle 3 추가분):")
        for bid, dn, bn, brand, matched_key, copies in new_variants_of_mapped:
            print(f"  id={bid:5d}  brand={brand}  matched={matched_key[:40]}")
            print(f"    host={bn[:90]}")
    if truly_new:
        print("\n[D] 완전히 새로운 케이스:")
        for bid, dn, bn, copies in truly_new:
            print(f"  id={bid:5d}  dn={dn[:25]!r}")
            print(f"    host={bn[:90]}")
            if copies:
                for cp in copies[:2]:
                    print(f"    copy: {cp[:80]}")


if __name__ == "__main__":
    main()
