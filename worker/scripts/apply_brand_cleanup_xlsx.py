"""Apply the user-edited brand-cleanup xlsx in one shot.

For every row whose '정정할 브랜드명' (column J) is filled, register the
brand's current host (a `__unverified__::<copy>` sentinel) in
canonical_brand.HOST_TO_BRAND. Then run reconcile_brands.py --apply so the
brands table's display_name catches up.

Run:
    uv run python scripts/apply_brand_cleanup_xlsx.py --dry-run
    uv run python scripts/apply_brand_cleanup_xlsx.py --apply
"""

from __future__ import annotations

import argparse
import re
import subprocess
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "브랜드정리_긴급정정_20260625_1039.xlsx"
CANON = (
    Path(__file__).resolve().parents[1]
    / "src" / "worker" / "lib" / "canonical_brand.py"
)


def read_xlsx() -> list[tuple[int, str, str]]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    out: list[tuple[int, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        brand_id = row[0]
        host = row[2]
        fix = row[9] if len(row) > 9 else None
        if not (brand_id and host and fix):
            continue
        fix = str(fix).strip()
        if not fix:
            continue
        out.append((int(brand_id), str(host).strip(), fix))
    return out


def insert_into_canonical(rows: list[tuple[int, str, str]], dry_run: bool) -> int:
    """Append new mappings to HOST_TO_BRAND in canonical_brand.py."""
    text = CANON.read_text(encoding="utf-8")

    # Find the existing __unverified__:: anchor (line 393 region) and append
    # a new dated batch immediately AFTER the existing 2026-05-27 batch's
    # closing line so the file stays grouped chronologically.
    anchor = '    "__unverified__::현대해상 굿앤굿어린이종합보험Q": "현대해상",\n'
    if anchor not in text:
        print("ERROR: anchor not found in canonical_brand.py")
        return 1

    kst = timezone(timedelta(hours=9))
    today = datetime.now(kst).strftime("%Y-%m-%d")
    batch_header = (
        f'    # ─── User-mapped {today} batch (긴급정정 xlsx) ─────────────\n'
    )

    # Collect existing keys so we don't add dupes
    existing_keys = set(re.findall(r'"(__unverified__::[^"]+)"\s*:', text))

    new_lines: list[str] = []
    added = []
    skipped_dup = []
    for brand_id, host, fix in rows:
        if not host.startswith("__unverified__::"):
            print(f"  [skip] id={brand_id} host not __unverified__::  ({host[:60]})")
            continue
        if host in existing_keys:
            skipped_dup.append((brand_id, host, fix))
            continue
        # escape backslash + quote inside the key
        safe_host = host.replace("\\", "\\\\").replace('"', '\\"')
        safe_brand = fix.replace("\\", "\\\\").replace('"', '\\"')
        new_lines.append(f'    "{safe_host}": "{safe_brand}",\n')
        added.append((brand_id, host, fix))

    insert_block = batch_header + "".join(new_lines)
    new_text = text.replace(anchor, anchor + insert_block, 1)

    print(f"  추가 {len(added)}건 / 기존 중복 skip {len(skipped_dup)}건")
    if skipped_dup:
        for r in skipped_dup[:5]:
            print(f"    [dup] {r[1][:70]} → {r[2]}")

    if not dry_run:
        CANON.write_text(new_text, encoding="utf-8")
        print(f"  [wrote] {CANON}")

    return 0


def run_reconcile(dry_run: bool) -> int:
    args = [
        "uv", "run", "python", "scripts/reconcile_brands.py",
    ]
    if not dry_run:
        args.append("--apply")
    print(f"\n>>> {' '.join(args)}")
    cp = subprocess.run(args, cwd=str(ROOT / "worker"))
    return cp.returncode


def main() -> int:
    p = argparse.ArgumentParser()
    grp = p.add_mutually_exclusive_group(required=True)
    grp.add_argument("--dry-run", action="store_true")
    grp.add_argument("--apply", action="store_true")
    a = p.parse_args()
    dry = not a.apply

    rows = read_xlsx()
    print(f"엑셀에서 정정 기재된 행: {len(rows)}건")
    rc = insert_into_canonical(rows, dry_run=dry)
    if rc:
        return rc
    return run_reconcile(dry_run=dry)


if __name__ == "__main__":
    sys.exit(main())
